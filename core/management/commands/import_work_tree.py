"""
Import work tree from structured Excel workbook "Структура ЕР.xlsx".

Builds parent-child relationships using the numbering system:
- Row 1: root (number 1.)
- Row 2: categories (numbers like 2.1, 2.2) — children of root
- Row 3: subcategories (numbers like 2.1.1, 2.1.2) — children of row 2 nodes
- Row 4+: deeper nodes (numbers 1., 2., 3., ... or 4. EL00-09-05, etc.) — 
  children of the nearest numbered node above with different number format
- Elements: codes without number prefix (like CS00-01-01-13) — 
  children of nearest grouping node in same column
- Variants: text without code or number (like 1-жильный) — grouping nodes
- Units: measurement keywords — attributes on elements, not tree nodes

Parent determination for numbered nodes:
  - Row 2 → always child of root
  - Row 3+ → look in same column for last numbered node with DIFFERENT format
    (single=N., double=N.N, triple=N.N.N...). Same format = sibling.
    If no node in same column, look across the band.
"""

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import WorkNode


HIERARCHY_PREFIX = re.compile(r"^\s*(\d+(?:\.\d+)*)\s*[\.\s)]")

CODE_PATTERN = re.compile(
    r"([A-ZА-ЯЁ]{2}[/A-ZА-ЯЁ]{0,3}(?:\d{2})?(?:-\d+)*)",
    re.IGNORECASE,
)

# Fallback for code prefixes that start with Latin letters but don't have
# the standard digit pattern (e.g. "EL/II Монтаж..." → code="EL/II")
LATIN_CODE_PREFIX = re.compile(r"^([A-Z]{2,}[\d/][A-Z\d/-]*)\s")

UNIT_KEYWORDS = re.compile(
    r"^(?:шт(?:ук[аи]?)?|"
    r"метры?|метр(?:ов)?|"
    r"тонны?|тонна|тонн|"
    r"м\b|м[234]|"
    r"кг|"
    r"килограмм[ыа]?|"
    r"100м(?:2)?|"
    r"10м[23]|"
    r"модуль|комплект|"
    r"т[\*\.]км|"
    r"усл\.\s*метр(?:ов)?|"
    r"лоток|"
    r"комплект)",
    re.IGNORECASE,
)

CODE_TRANSLATION = str.maketrans({
    "А": "A", "В": "B", "С": "C", "Е": "E", "Н": "H",
    "К": "K", "М": "M", "О": "O", "Р": "P", "Т": "T",
    "Х": "X", "У": "Y", "І": "I", "Ї": "I", "Ё": "E",
    "а": "A", "в": "B", "с": "C", "е": "E", "н": "H",
    "к": "K", "м": "M", "о": "O", "р": "P", "т": "T",
    "х": "X", "у": "Y", "і": "I", "ї": "I", "ё": "E",
    "/": " ",
})


def normalize_code(code: str) -> str:
    return code.translate(CODE_TRANSLATION).upper()


def clean(text: str) -> str:
    if not text:
        return ""
    t = str(text).replace("\r", " ").replace("\n", " ").replace("\xa0", " ")
    t = re.sub(r"\s+", " ", t).strip()
    t = re.sub(r"^[\s.\-–—:]+", "", t)
    return t.strip()


def extract_number_and_rest(text: str) -> tuple[str, str]:
    m = HIERARCHY_PREFIX.match(text)
    if m:
        number = m.group(1)
        rest = text[m.end():].strip()
        return number, rest
    return "", text.strip()


def extract_code(text: str) -> tuple[str, str]:
    m = CODE_PATTERN.search(text)
    if m:
        code = normalize_code(m.group(1))
        if re.search(r"\d{2,}", code):
            name = text[m.end():].strip()
            return code, name or code

    # Fallback: text starts with a Latin code prefix (e.g. "EL/II Монтаж...")
    m2 = LATIN_CODE_PREFIX.match(text)
    if m2:
        code = normalize_code(m2.group(1))
        name = text[m2.end():].strip()
        return code, name or code

    return "", text.strip()


def number_format(number: str) -> str:
    dots = number.count(".")
    if dots == 0:
        return "single"
    elif dots == 1:
        return "double"
    else:
        return "triple"


def is_unit(text: str) -> bool:
    if not text or len(text) > 60:
        return False
    # Only skip if text contains a real code with digits (not pure Cyrillic like "метр")
    m = CODE_PATTERN.search(text)
    if m and re.search(r"\d{2,}", m.group(1)):
        return False
    if LATIN_CODE_PREFIX.match(text):
        return False
    t = text.strip().split("(")[0].split(" (")[0].strip()
    return bool(UNIT_KEYWORDS.match(t))


def is_variant(text: str) -> bool:
    t = text.strip()
    if not t or len(t) > 60:
        return False
    # A real code must have 2+ digits (to not match pure Cyrillic words like "жильный")
    m = CODE_PATTERN.search(t)
    if m and re.search(r"\d{2,}", m.group(1)):
        return False
    if LATIN_CODE_PREFIX.match(t):
        return False
    if HIERARCHY_PREFIX.match(t):
        return False
    if is_unit(t):
        return False
    if re.search(r"\d.*жил", t, re.IGNORECASE):
        return True
    return False


class Command(BaseCommand):
    help = "Import work tree from structured Excel workbook"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--clear", action="store_true",
                            help="Deactivate all existing nodes before import")
        parser.add_argument("--verbose", action="store_true",
                            help="Print detailed info for each node")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        workbook = load_workbook(file_path, data_only=True)
        if not workbook.worksheets:
            raise CommandError("No sheets in workbook.")

        verbose = options["verbose"]

        if options["clear"]:
            deactivated = WorkNode.objects.update(is_active=False)
            self.stdout.write(f"Deactivated: {deactivated}")

        total_c = total_u = 0

        with transaction.atomic():
            for sheet in workbook.worksheets:
                self.stdout.write(f"\n{'=' * 60}")
                self.stdout.write(f"Sheet: {sheet.title}")
                c, u = self.import_sheet(sheet, verbose)
                total_c += c
                total_u += u

            cleared = self.clear_units_from_parents()
            if cleared:
                self.stdout.write(f"Cleared units from {cleared} parent nodes")

        self.stdout.write(
            self.style.SUCCESS(f"\nDone. Created: {total_c}, Updated: {total_u}")
        )

    def import_sheet(self, sheet, verbose=False):
        grid = self._read_grid(sheet)

        root_text = grid.get((1, 1), "")
        if not root_text:
            self.stdout.write(self.style.WARNING("  Skipped: no A1"))
            return 0, 0

        root_number, root_rest = extract_number_and_rest(root_text)
        root_code, root_name = extract_code(root_rest)
        if not root_code:
            root_code = self._sheet_code(sheet.title)
        if not root_name:
            root_name = root_rest or root_text

        root_node, _ = self._upsert(
            f"{sheet.title}:A1", root_code, root_name, None,
        )
        self.stdout.write(f"  Root: {root_code} | {root_name[:50]}")

        band_starts = sorted({col for (r, col), _ in grid.items() if r == 2})
        if not band_starts:
            band_starts = [1]

        bands = {}
        for bs in band_starts:
            bands[bs] = {
                "unit": "",
                "all_nodes": [],
                "col_nodes": {},
                "col_last_grouping": {},
                "col_had_elements": {},
            }

        created = updated = 0
        max_row = max((r for r, _ in grid), default=1)

        for row_num in range(2, max_row + 1):
            cols_in_row = sorted(c for r, c in grid if r == row_num)
            for col in cols_in_row:
                text = grid[(row_num, col)]
                if not text:
                    continue

                band = self._resolve_band(col, band_starts)
                bd = bands[band]

                if col not in bd["col_nodes"]:
                    bd["col_nodes"][col] = []
                if col not in bd["col_last_grouping"]:
                    bd["col_last_grouping"][col] = root_node
                if col not in bd["col_had_elements"]:
                    bd["col_had_elements"][col] = False

                if is_unit(text):
                    # Extract just the unit keyword, stripping parenthetical notes
                    bd["unit"] = text.strip().split("(")[0].split(" (")[0].strip()
                    if verbose:
                        self.stdout.write(
                            f"    [R{row_num}C{col:2d}] UNIT: {text[:30]}"
                        )
                    continue

                number, rest = extract_number_and_rest(text)

                if number:
                    fmt = number_format(number)

                    if row_num == 2:
                        parent = root_node
                    else:
                        parent = self._find_parent_for_numbered(
                            number, fmt, col, bd, root_node, row_num
                        )

                    bd["col_had_elements"][col] = False

                    code, name = extract_code(rest)
                    if not code:
                        code = ""
                    if not name:
                        name = rest

                    node, was_new = self._upsert(
                        f"{sheet.title}:R{row_num}C{col}",
                        code, name, parent, bd["unit"],
                    )
                    created += int(was_new)
                    updated += int(not was_new)

                    bd["all_nodes"].append((row_num, col, fmt, node))
                    bd["col_nodes"][col].append((row_num, fmt, node))
                    bd["col_last_grouping"][col] = node

                    if verbose:
                        prefix = "  " * len(bd["all_nodes"])
                        self.stdout.write(
                            f"    [R{row_num}C{col:2d}] {number:15s} {prefix}"
                            f"{code or '—':20s} {name[:30]}"
                            + (f" [{bd['unit']}]" if bd["unit"] else "")
                        )

                elif is_variant(text):
                    parent = self._find_parent_for_variant(col, bd, root_node)
                    node, was_new = self._upsert(
                        f"{sheet.title}:R{row_num}C{col}",
                        "", text, parent, "",
                    )
                    created += int(was_new)
                    updated += int(not was_new)
                    bd["col_last_grouping"][col] = node
                    bd["col_had_elements"][col] = False

                    if verbose:
                        self.stdout.write(
                            f"    [R{row_num}C{col:2d}] VARIANT: {text[:40]}"
                        )

                else:
                    code, name = extract_code(text)
                    if code:
                        parent = bd["col_last_grouping"].get(col, root_node)
                        node, was_new = self._upsert(
                            f"{sheet.title}:R{row_num}C{col}",
                            code, name or text, parent, bd["unit"],
                        )
                        created += int(was_new)
                        updated += int(not was_new)

                        bd["col_had_elements"][col] = True

                        if verbose:
                            self.stdout.write(
                                f"    [R{row_num}C{col:2d}] ELEMENT: "
                                f"{code:20s} {name[:30]}"
                                + (f" [{bd['unit']}]" if bd["unit"] else "")
                            )
                    else:
                        if verbose:
                            self.stdout.write(
                                f"    [R{row_num}C{col:2d}] SKIP: {text[:40]}"
                            )

        self.stdout.write(f"  Created: {created}, Updated: {updated}")
        return created, updated

    def _find_parent_for_numbered(self, number, fmt, col, bd, root_node, row_num):
        col_nodes = bd["col_nodes"].get(col, [])

        if col_nodes:
            last_r, last_fmt, last_node = col_nodes[-1]

            if last_fmt != fmt:
                # Different format → parent is the last node in this column
                return last_node

            # Same format → check had_elements
            had_elems = bd["col_had_elements"].get(col, False)
            if not had_elems:
                # No elements since last grouping → child of last node
                return last_node
            else:
                # Elements since last grouping → sibling
                # Find the sibling's parent (previous node with different format in same column)
                for r, f, n in reversed(col_nodes[:-1]):
                    if f != fmt:
                        return n
                return root_node

        # First numbered node in this column – find nearest parent in band
        for r, c, nfmt, n in reversed(bd["all_nodes"]):
            if r >= row_num:
                continue
            if nfmt != fmt:
                return n

        return root_node

    def _find_parent_for_variant(self, col, bd, root_node):
        col_nodes = bd["col_nodes"].get(col, [])
        if col_nodes:
            return col_nodes[-1][2]
        if bd["all_nodes"]:
            return bd["all_nodes"][-1][3]
        return root_node

    def _read_grid(self, sheet):
        grid = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val = clean(str(cell.value))
                    if val:
                        grid[(cell.row, cell.column)] = val
        return grid

    def _resolve_band(self, col, band_starts):
        best = None
        for s in band_starts:
            if s <= col:
                best = s
            else:
                break
        return best

    def _upsert(self, source_key, code, name, parent, unit=""):
        return WorkNode.objects.update_or_create(
            source_key=source_key,
            defaults={
                "code": code,
                "name": name,
                "parent": parent,
                "unit": unit,
                "is_active": True,
            },
        )

    def _sheet_code(self, title):
        m = CODE_PATTERN.match(title)
        if m and re.search(r"\d{2,}", m.group(1)):
            return normalize_code(m.group(1))
        m2 = LATIN_CODE_PREFIX.match(title)
        if m2:
            return normalize_code(m2.group(1))
        return title.split()[0] if title.split() else title[:10]

    def clear_units_from_parents(self):
        parent_ids = WorkNode.objects.filter(
            is_active=True,
            children__is_active=True,
        ).values_list("id", flat=True).distinct()
        return WorkNode.objects.filter(
            id__in=list(parent_ids),
        ).exclude(unit="").update(unit="")
