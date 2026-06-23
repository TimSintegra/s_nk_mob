from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Brigade, Worker


class Command(BaseCommand):
    help = "Import workers from СМУ.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--brigade", type=str, help="Название бригады для привязки рабочих")
        parser.add_argument("--dry-run", action="store_true", help="Показать что будет импортировано без сохранения")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"Файл не найден: {file_path}")

        workbook = load_workbook(file_path, data_only=True)
        ws = workbook.active

        brigade = None
        if options["brigade"]:
            brigade = Brigade.objects.filter(name=options["brigade"]).first()
            if not brigade:
                self.stdout.write(self.style.WARNING(f"Бригада '{options['brigade']}' не найдена. Рабочие будут без бригады."))

        created_count = 0
        skipped_count = 0
        results = []

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
                full_name = str(row[0]).strip() if row[0] else ""
                if not full_name:
                    continue

                existing = Worker.objects.filter(full_name=full_name).first()
                if existing:
                    skipped_count += 1
                    results.append((full_name, "УЖЕ ЕСТЬ"))
                    continue

                if not options["dry_run"]:
                    Worker.objects.create(
                        full_name=full_name,
                        brigade=brigade,
                        is_active=True,
                    )

                created_count += 1
                results.append((full_name, "СОЗДАН"))

        self.stdout.write("")
        self.stdout.write(f"Создано: {created_count}")
        self.stdout.write(f"Пропущено (уже есть): {skipped_count}")

        if results:
            self.stdout.write("\nДетали:")
            for name, status in results:
                self.stdout.write(f"  {name:40s} [{status}]")

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Сухой прогон — изменения не сохранены."))
