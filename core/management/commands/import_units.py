from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import WorkNode


class Command(BaseCommand):
    help = "Import units of measurement from ЕР.xlsx into WorkNode records"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--dry-run", action="store_true", help="Only show what would change")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"Файл не найден: {file_path}")

        workbook = load_workbook(file_path, data_only=True)
        if not workbook.worksheets:
            raise CommandError("В книге нет листов.")

        sheet = workbook.worksheets[0]
        self.stdout.write(f"Лист: '{sheet.title}' (строк: {sheet.max_row})")

        updated_count = 0
        not_found_count = 0
        skipped_count = 0
        results = []

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = str(row[2]).strip() if row[2] else ""  # column C
                unit = str(row[8]).strip() if row[8] else ""   # column I (unit)
                name = str(row[7]).strip() if row[7] else ""   # column H (name)

                if not code:
                    continue

                # Normalize unit — remove empty/None
                if not unit or unit.lower() in ("none", "null", ""):
                    skipped_count += 1
                    continue

                try:
                    node = WorkNode.objects.get(code=code, is_active=True)
                except WorkNode.DoesNotExist:
                    not_found_count += 1
                    results.append((code, unit, name, "НЕ НАЙДЕН"))
                    continue
                except WorkNode.MultipleObjectsReturned:
                    # If multiple nodes match, try the leaf-most one
                    node = WorkNode.objects.filter(code=code, is_active=True).last()

                old_unit = node.unit
                if old_unit == unit:
                    skipped_count += 1
                    continue

                if not options["dry_run"]:
                    node.unit = unit
                    node.save(update_fields=["unit"])

                updated_count += 1
                results.append(
                    (code, unit, name[:50], f"{old_unit or '—'} → {unit}")
                )

        # Print results
        self.stdout.write("")
        if updated_count:
            self.stdout.write(f"Обновлено: {updated_count}")
        if not_found_count:
            self.stdout.write(self.style.WARNING(f"Не найдено в БД: {not_found_count}"))
        if skipped_count:
            self.stdout.write(f"Пропущено (нет единицы/совпадает): {skipped_count}")

        if results:
            self.stdout.write("\nДетали:")
            for code, unit, name, status in results[:50]:
                self.stdout.write(f"  {code:25s} → {unit:10s} [{status}]")
            if len(results) > 50:
                self.stdout.write(f"  ... и ещё {len(results) - 50}")

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Сухой прогон — изменения не сохранены."))
